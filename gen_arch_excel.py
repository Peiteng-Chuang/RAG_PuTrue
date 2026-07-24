# -*- coding: utf-8 -*-
"""產出 RAG 系統端到端架構 Excel（資料輸入 → ETL → 向量化/儲存 → 檢索 → LLM → UI 回應）。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 樣式 ----
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="Microsoft JhengHei", color="FFFFFF", bold=True, size=11)
ETL_FILL = PatternFill("solid", fgColor="DDEBF7")   # 淺藍 = ETL
LLM_FILL = PatternFill("solid", fgColor="FCE4D6")   # 淺橘 = 檢索/LLM
CELL_FONT = Font(name="Microsoft JhengHei", size=10)
SECT_FONT = Font(name="Microsoft JhengHei", size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_rows(ws, rows, widths, fill_col=None):
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = TOP
            cell.border = BORDER
        # 依「類別」欄上色（fill_col = 0-based index of category col）
        if fill_col is not None:
            cat = row[fill_col]
            fill = ETL_FILL if "ETL" in cat or "前處理" in cat else (
                LLM_FILL if ("檢索" in cat or "LLM" in cat or "UI" in cat) else None)
            if fill:
                for c_idx in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


wb = Workbook()

# ========== Sheet 1：端到端流程（主表）==========
ws1 = wb.active
ws1.title = "端到端流程"
cols1 = ["#", "管線分段", "階段", "使用技術 / 工具", "角色 / 做了什麼", "執行環境", "備註 / 關鍵設計", "優化方向（ROI: 🔴高 🟡中 🟢低）"]
ws1.append(cols1)
style_header(ws1, len(cols1))

p1 = [
    [1, "① ETL 前處理", "資料來源輸入",
     "原始檔（PDF / PPT / Excel / DOC）",
     "1TB 建築工程文件：CAD 圖、表格、施工規範、術語",
     "DATA_ROOT (D:/璞真RAG資料夾/)",
     "依資料夾分類：doc_type=類別、project_name=建案（schema 2.1.1）",
     "🟡 統一 data_root 來源：run_v5(--data-root 預設 璞真RAG_rawdata)、rebuild_tracker(璞真RAG資料夾)、.env(D:/璞真RAG資料夾) 三處預設不一致，忘帶參數會切錯 file_key。建議集中讀 .env DATA_ROOT。🟡 ingest 前先建 file inventory + hash 去重"],
    [2, "① ETL 前處理", "格式預轉換",
     "LibreOffice (headless)",
     "PPT/DOC → PDF，統一成 PDF 入口",
     "4090 批次端",
     "ppt_to_pdf.py；曾踩 LO timeout 已修",
     "🟡 LibreOffice 單檔序列轉換是批次瓶頸 → 轉檔池化/並行（已有 per-file profile 隔離，可安全並行）"],
    [3, "① ETL 前處理", "自動分流 Triage",
     "PyMuPDF (fitz) + 自訂規則",
     "page.get_drawings() > 200 → 判定 CAD/複雜頁，切慢路徑",
     "4090 批次端",
     "200 向量閾值＝成本/精度交界（CAD 結構符號每個上百向量）",
     "🔴 加 page.find_tables() OR 條件：純表格頁 drawings 常 <30→誤走快路徑丟 cell 結構（最小改動高 ROI，仍 TODO）。🟡 default 仍 DrawingCountTriage；已寫好的 MultiSignal/Adaptive 在 4090 跑閾值對比（200/300）後評估換 default"],
    [4, "① ETL 前處理", "文字抽取（快路徑）",
     "PyMuPDF",
     "純文字/簡單頁直接抽 text + span/bbox",
     "4090 批次端",
     "快；多數頁走此路",
     "🟡 多欄閱讀順序錯亂：fast path 用 y0 排序，左右雙欄會交錯 → 加欄位偵測 / x-cut 分欄再排序"],
    [5, "① ETL 前處理", "結構/OCR 還原（慢路徑）",
     "Marker + Surya + Nougat（GPU）",
     "CAD/複雜表格/掃描頁做 layout 分析 + OCR",
     "4090 GPU",
     "非對稱雙路：慢但精；v5 入口須 top-level import torch 避免誤判 hang",
     "🟡 v5 比 v4 慢 1.5-3x：marker_pool 已寫但待 4090 驗 VRAM/加速比；torch.compile 仍 TODO。🟡 Marker per-page timeout（R8 deferred）防單頁 hang 拖垮整批"],
    [6, "① ETL 前處理", "結構化還原",
     "Regex + Python multiprocessing",
     "重建標題階層、頁碼剃除、垃圾/浮水印清除",
     "4090 批次端",
     "四層階層：# stem / ## 第N頁 / ### page_title / #### subtitle",
     "🟡 無抽取品質閘門：加品質訊號（空 body / mojibake / 多欄錯序偵測），標記壞頁讓 review 優先看，而非靠人逐頁翻"],
    [7, "① ETL 前處理", "Citation 錨點",
     "自訂 Markdown 慣例",
     "每頁插入 `## 第N頁` 作為頁級來源追溯錨點",
     "—",
     "結構性 regex 嚴守原樣 + UI 防呆（核心溯源價值）",
     "🟢 已穩固。可加錨點完整性自動驗證：每頁應有且唯一 `## 第N頁`，缺漏/重複即告警（防下游 page_idx 漂移）"],
    [8, "① ETL 前處理", "圖片抽取 / 多模態關聯",
     "PyMuPDF image extract",
     "抽頁面圖、與文字頁強關聯，供雙 collection 用",
     "4090 批次端",
     "向量圖表流失為已知技術債",
     "🟡 向量繪製圖表流失：fast path 只抓點陣圖(get_images)，<200 向量的向量圖整個不見 → 對「中量向量頁」rasterize（需你拍板，避免裝飾向量變雜訊圖）。🟡 接 P6 圖片 collection 才能真正多模態檢索"],
    [9, "① ETL 前處理", "產出物",
     "Markdown + sidecar JSON",
     "一檔一份 .md + 對應 metadata JSON（payload 來源）",
     "—",
     "詳見 qdrant格式.md；point_id = file_hash",
     "🟡 孤兒清理：raw 刪檔後 .md/sidecar/向量/points 全殘留無 cascade → 加 orphan GC 工具（對照實體目錄掃殘留）"],
    [10, "② Review/標記", "人工 review / 標記",
     "Streamlit (md_review_ui.py, 4426 行)",
     "PDF 原圖 vs Markdown 並排、doc_type 一鍵標記、chunk preview",
     "2050 互動端",
     "ETL 品質把關駕駛艙；downstream 工具，與 ETL upstream 隔離",
     "🟡 跨頁編輯未 commit 到 current_md 的潛在 bug（翻頁前先 commit 舊頁，記憶已記備修）。🟡 4426 行單檔可考慮拆模組降回歸風險。🟡 對外服務化走 Dify External Knowledge API（不互通 DB）"],
    [11, "③ 向量化/儲存", "Chunking",
     "自訂 heading+token-v2",
     "依標題切塊，heading 階層當 breadcrumb metadata",
     "2050 互動端",
     "孤兒葉標題降級，避免整頁皆標題→0 chunk 資訊遺失",
     "🔴 chunk 無長度截斷 + top_k=8 → 易超出 qwen3:32b ~3000字注意力 → 加 chunk 截斷(400-600字/筆)或總長 guard（便宜、直接提升答案品質）"],
    [12, "③ 向量化/儲存", "Embedding 向量化",
     "bge-m3 (hybrid: dense + sparse named vectors)",
     "同時產 dense 與 sparse 向量",
     "2050 GPU（baseline 須 4090）",
     "GPU 須主執行緒載入/推理，嚴禁背景執行緒載入",
     "🟡 2050 4GB 互動端載 bge-m3 有 OOM/阻塞風險 → 把 embedding 抽成獨立行程/HTTP 服務（解 Streamlit 主執行緒阻塞 + 可跨機共用，記憶指定的非阻塞正解之一）"],
    [13, "③ 向量化/儲存", "向量儲存",
     "Qdrant (Docker server)",
     "雙 collection：text + images 分庫；payload 存來源/頁/doc_type",
     "2050 互動端 (Docker)",
     "payload-only 遷移可免重 embed（set_payload）",
     "🟡 孤兒 point 無人清：內容改→新 point_id，舊 points 留庫稀釋檢索 → 加 GC（依 file_hash 對照清舊版本）。🟢 stem 撞號目前只有 rebuild_tracker 擋，可前移到 ingest 偵測"],
    [14, "④ 檢索 RAG", "Query 改寫 (HyDE)",
     "vLLM (qwen3:32b)",
     "送檢索前用 LLM 生成假設答案再 encode，提升召回",
     "vLLM 服務 (192.168.201.66:8000)",
     "自建有 HyDE，Dify 無原生需自組",
     "🟢 HyDE 也吃一次 LLM 延遲 → 待 eval set 建好後 A/B 量化 on/off 對 recall 的實際增益，無效則對短 query 關閉省延遲"],
    [15, "④ 檢索 RAG", "過濾 (Filter)",
     "Qdrant filter + facet_filter",
     "doc_type / project_name 連動過濾（專案連動於已選類別）",
     "2050 互動端",
     "兩條 must（條內 OR、條間 AND），杜絕空組合",
     "🟢 連動過濾已完善。可擴維度：年份/檔案類型/review_status，並讓檢索只吃 ingested 狀態的 point"],
    [16, "④ 檢索 RAG", "檢索",
     "Qdrant search (dense / sparse / hybrid 三選)",
     "向量相似檢索，top_k=8",
     "2050 互動端",
     "⚠️ 無 rerank 階段（已知缺口）；無 chunk 截斷有稀釋風險",
     "🔴 最大缺口：加 rerank（bge-reranker-v2-m3 同家族）→ 先大 top_k 召回再 rerank 收斂到 3-5 筆，召回率+精度雙升，是補齊 vs Dify 的關鍵。🟡 調 hybrid dense/sparse fusion 權重"],
    [17, "⑤ LLM 生成", "答案生成",
     "vLLM (OpenAI 相容 /v1) + qwen3:32b",
     "把 query + 檢索 chunks 組 prompt 生成回答",
     "vLLM 服務",
     "已取代 Ollama；num_ctx 由服務端 --max-model-len 決定",
     "🟡 套用「提示-內文-提示」prompt 架構（pipeline.ipynb hint，緩解長 context 稀釋）。🟡 強制 inline citation/structured output，讓回答可追回頁錨點"],
    [18, "⑤ LLM 生成", "LLM 開關",
     "UI toggle（🤖 LLM）",
     "關＝僅檢索模式，不呼叫 LLM，只回「N 個相關文件」+ 來源",
     "2050 互動端",
     "送出箭頭旁開關",
     "🟢 可加「自動模式」：依檢索最高分門檻決定要不要呼叫 LLM（低分直接回來源，省 token 與延遲）"],
    [19, "⑥ UI 回應", "對話記憶",
     "streamlit-local-storage (瀏覽器 localStorage)",
     "近 5 輪記憶存瀏覽器端，不寫 server 磁碟",
     "瀏覽器",
     "setItem 後勿緊接 st.rerun()；persist 延到 render 尾端",
     "🟢 長對話只留近 5 輪會丟前文 → 加滾動式 conversation summary 記憶。🟢 localStorage 限本瀏覽器、不跨裝置（可接受，記下即可）"],
    [20, "⑥ UI 回應", "回應呈現",
     "Streamlit chat view",
     "回答黏 chunks（檔名/score/🔗link/📂開檔/卡牌頁圖 dialog）+ 原地內嵌 PNG",
     "2050 互動端",
     "頁級溯源 UX，為 review 場景特製",
     "🟢 來源 UX 已是強項。可加：回答內 inline 引用標記直接連回對應頁錨點/頁圖，強化逐句溯源"],
]
write_rows(ws1, p1, [4, 12, 14, 24, 30, 18, 32, 52], fill_col=1)
ws1.row_dimensions[1].height = 24

# ========== Sheet 2：技術棧總覽（依層分類）==========
ws2 = wb.create_sheet("技術棧總覽")
cols2 = ["層級", "技術 / 套件", "版本 / 設定", "用途"]
ws2.append(cols2)
style_header(ws2, len(cols2))
p2 = [
    ["執行環境", "Python venv (venv312)", "Python 3.12", "全系統執行環境"],
    ["執行環境", "PyTorch + CUDA", "GPU 加速", "Marker/Surya/Nougat、bge-m3 推理"],
    ["ETL 抽取", "PyMuPDF (fitz)", "快路徑", "文字/向量/圖片抽取、triage 判定"],
    ["ETL 抽取", "Marker / Surya / Nougat", "慢路徑 GPU", "layout 分析 + OCR（CAD/表格/掃描）"],
    ["ETL 抽取", "LibreOffice (headless)", "—", "PPT/DOC → PDF 預轉換"],
    ["ETL 抽取", "Poppler", "25.12.0", "PDF 工具（POPPLER_PATH）"],
    ["ETL 處理", "Regex + multiprocessing", "—", "結構化還原、標題階層、清洗、並行"],
    ["Embedding", "bge-m3", "hybrid dense+sparse", "向量化（named vectors）"],
    ["向量庫", "Qdrant", "Docker server", "雙 collection（text+images）、payload 過濾"],
    ["LLM 後端", "vLLM (OpenAI 相容)", "base_url 含 /v1 :8000", "HyDE / 標題 / 主回答；取代 Ollama"],
    ["LLM 模型", "qwen3:32b", "LLM_MODEL_NAME", "主對話 LLM（model id 由 /v1/models 動態抓）"],
    ["LLM 模型", "qwen3-vl:latest", "VLM_MODEL_NAME", "視覺語言模型（fallback 設定）"],
    ["LLM 客戶端", "openai (python)", "—", "呼叫 vLLM OpenAI 相容端點"],
    ["前端 UI", "Streamlit", "layout=wide", "md_review_ui.py：review + 對話 + Qdrant 維運"],
    ["前端 UI", "streamlit-local-storage", "—", "對話記憶存瀏覽器 localStorage"],
    ["輔助", "Gemini API", "GEMINI_API_KEY", "（輔助/實驗用）"],
]
for r_idx, row in enumerate(p2, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = TOP
        cell.border = BORDER
    layer = row[0]
    fill = ETL_FILL if layer.startswith("ETL") or layer in ("執行環境",) else (
        LLM_FILL if layer.startswith("LLM") or layer in ("前端 UI", "向量庫", "Embedding") else None)
    if fill:
        for c_idx in range(1, len(row) + 1):
            ws2.cell(row=r_idx, column=c_idx).fill = fill
for i, w in enumerate([14, 26, 22, 46], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ========== Sheet 3：硬體環境 ==========
ws3 = wb.create_sheet("硬體環境")
cols3 = ["機器", "角色", "負責階段", "備註"]
ws3.append(cols3)
style_header(ws3, len(cols3))
p3 = [
    ["RTX 2050 機", "互動端", "md_review_UI / Qdrant / 向量化 / 檢索 / 對話", "日常 review 與問答"],
    ["RTX 4090 機", "批次端", "v5 ETL / Marker GPU 慢路徑 / OCR", "baseline 與大批 ETL 必須在此跑"],
    ["vLLM 服務", "LLM 推理", "HyDE / 標題 / 主回答生成", "192.168.201.66:8000，OpenAI 相容 /v1"],
]
for r_idx, row in enumerate(p3, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = TOP
        cell.border = BORDER
for i, w in enumerate([16, 14, 40, 36], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ========== Sheet 4：優化優先序 roadmap ==========
ws4 = wb.create_sheet("優化優先序")
cols4 = ["優先序", "對應階段", "優化項目", "為什麼（痛點）", "預估成本 / 備註"]
ws4.append(cols4)
style_header(ws4, len(cols4))
p4 = [
    ["🔴 P1", "（跨階段）", "建立評估集 + recall@K 數字化（P7）",
     "目前所有優化都靠手感判斷，無法量測哪個真有效；這是其他優化的前提",
     "中：標 30-50 題 query→正解 chunk 的 golden set，跑 recall@K / MRR"],
    ["🔴 P2", "⑯ 檢索", "加 rerank（bge-reranker-v2-m3）",
     "最大檢索缺口；大 top_k 召回後 rerank 收斂，召回+精度雙升，補齊 vs Dify",
     "低-中：同家族模型、現成；2050 顯存需評估或放 4090/獨立服務"],
    ["🔴 P3", "⑪ Chunking / ⑯ 檢索", "chunk 長度截斷 / 總長 guard",
     "top_k=8 無截斷易超出 qwen3:32b ~3000字注意力，答案被稀釋",
     "低：在 llm_format_chunks 外套截斷(400-600字/筆)，當天可改"],
    ["🔴 P4", "③ Triage", "page.find_tables() OR 條件",
     "純表格頁 drawings<30 誤走快路徑、丟 cell 結構（建築表格是核心內容）",
     "低：triage 加一個 OR 訊號，記憶標為最小改動高 ROI"],
    ["🟡 P5", "⑬ 儲存 / ⑨ 產出", "孤兒 point / 檔案 cascade GC",
     "內容改→新 point_id 舊的殘留稀釋檢索；raw 刪檔殘留無人清",
     "中：寫 GC 工具，依 file_hash 對照實體目錄清殘留"],
    ["🟡 P6", "⑫ Embedding", "embedding 抽成獨立行程 / 服務",
     "2050 4GB 載 bge-m3 阻塞主執行緒、有 OOM 風險；無法跨機共用",
     "中：subprocess 或 HTTP 服務，記憶指定的非阻塞正解之一"],
    ["🟡 P7", "⑤ 慢路徑", "v5 速度（marker_pool / torch.compile）+ 4090 驗證",
     "v5 比 v4 慢 1.5-3x；pool 已寫但未在 4090 驗 VRAM/加速比",
     "中：需 4090 session 跑 baseline 與閾值對比實驗"],
    ["🟡 P8", "④/⑥ 抽取", "多欄排序修正 + 抽取品質閘門",
     "雙欄 y0 排序交錯；無訊號標記抽壞頁",
     "中：欄位偵測 x-cut + 品質訊號（空body/mojibake/錯序）"],
    ["🟢 P9", "①/⑩ 工程", "統一 data_root 來源 + 拆 4426 行單檔",
     "三處 data_root 預設不一致易切錯 file_key；單檔維護面積大",
     "低：集中讀 .env DATA_ROOT；UI 漸進拆模組"],
]
for r_idx, row in enumerate(p4, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = TOP
        cell.border = BORDER
    tag = row[0]
    fill = PatternFill("solid", fgColor="FFC7CE") if "🔴" in tag else (
        PatternFill("solid", fgColor="FFEB9C") if "🟡" in tag else
        PatternFill("solid", fgColor="C6EFCE"))
    ws4.cell(row=r_idx, column=1).fill = fill
for i, w in enumerate([8, 16, 30, 40, 38], start=1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

out = r"C:\project_file\RAG_PuTrue\RAG系統架構_技術盤點.xlsx"
wb.save(out)
print("saved:", out)
